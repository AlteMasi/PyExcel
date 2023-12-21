import openpyxl
from openpyxl import Workbook
from typing import Union

class Cell:
    def __init__(self, x:Union[str,int] = 1, y:int = 1):
        self.x = x
        self.y = y

    def get_letter(self, number):
        delta = 64
        if number > 90 - delta: 
            prefix = chr(int(number/26)+delta)
            if ord(prefix) > 90: return None
            return prefix + chr(number%26+delta+1)
        return chr(number+delta)

    def convert(self)->None:
        if type(self.x) is int: self.x = self.get_letter(self.x)

class ExcelFile:
    def __init__(self, name:str="excelfile", path:str=None) -> None:
        """ 
            Create an Excel File.

            Example:
            ```
            file = ExcelFile(name=filename)
            file.write(Cell("A",1),value,False)
            print(file.read(Cell(1,1)))
            file.save() 
            ```
        """
        self.name = name
        self.__extension = ".xlsx"
        self.path = path if path else f"{self.name}{self.__extension}"
        self.open()
        self.update()
    
    def open(self):
        """ Open Workbook """
        # try to load workbook
        self.wb = "workbook"
        try: self.wb = openpyxl.load_workbook(self.path)
        # create a new workbook the current one doesn't exist
        except FileNotFoundError: self.wb = Workbook()
        self.ws = self.wb.active
    
    def save(self):
        """ Save Workbook """
        try: self.wb.save(self.path)
        except PermissionError: raise PermissionError("can't save the file because it's already open")

    def change_sheet(self, name):
        """ change worksheet """
        self.ws = self.wb[name]

    def create_sheet(self, name):
        """ create a worksheet """
        self.ws = self.wb.create_sheet(name)
        self.save()

    def change_sheet_name(self, name):
        """ change a worksheet name """
        self.ws.title = name
        self.save()
    
    def write(self, cell:Union[Cell,tuple], value:str, save:bool=True):
        """ write value on a cell object """
        # convert cell to Cell
        cell : Cell = self.get_cell(cell)
        # if x is number -> convert to letter
        cell.convert()
        # write value on the cell
        self.ws[f"{cell.x}{cell.y}"] = value
        # if save is True then save Workbook
        if save: self.save()

    def write_range(self, row:Union[None,str,int]=None, col:Union[None,str,int]=None, values:list[str]=[], save:bool=True):
        """ write values in a cell range """
        if not row and not col: return None
        elif not values: return None
        cell_number =  self.max_column if row else self.max_row
        while True:
            if len(values) == cell_number: break
            elif len(values) > cell_number: 
                cell_number=len(values)
                break
            values.append("")
        for index in range(1,cell_number+1): 
            cell = (row,index) if row else (index,col)
            self.ws.cell(row=cell[0],column=cell[1]).value = values[index-1]
        # if save is True then save Workbook
        if save: self.save()

    def read(self, cell:Union[Cell,tuple])->str:
        """ read value of a cell object """
        # convert cell to Cell
        cell : Cell = self.get_cell(cell)
        # if x is number -> convert to letter
        cell.convert()
        # read value of a cell
        return self.ws[f"{cell.x}{cell.y}"].value

    def read_range(self, row:Union[None,str,int]=None, col:Union[None,str,int]=None, clear_empty:bool=True)->list:
        """ read values in a cell range """
        values = []
        if not row and not col: return None
        cell_number =  self.max_column if row else self.max_row
        for index in range(1,cell_number+1): 
            cell = (row,index) if row else (index,col)
            value = self.ws.cell(row=cell[0],column=cell[1]).value
            if clear_empty and value == None: continue
            values.append(value)
        return values

    def get_cell(self, cell:Union[Cell,tuple])->Cell:
        if type(cell) in [tuple,list]:
            return Cell(cell[0],cell[1])
        elif type(cell) is Cell:
            return cell
        else:
            raise TypeError(f"{type(cell)} should be a Cell, a tuple or a list")

    def __get_max_row(self):
        self.max_row = self.ws.max_row

    def __get_max_column(self):
        self.max_column = self.ws.max_column

    def update(self):
        self.__get_max_row()
        self.__get_max_column()